from flask import Flask, render_template, request, redirect, url_for, jsonify
import openpyxl
import os
from datetime import datetime
import pandas as pd
import json

app = Flask(__name__)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "12345"

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            return redirect(url_for('dashboard'))
        else:
            error = "Invalid username or password"
    return render_template('login.html', error=error)

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/scan', methods=['POST'])
def scan_qr():
    try:
        qr_data = request.get_json()

        roll = qr_data.get("roll")
        name = qr_data.get("name")
        student_class = qr_data.get("class")
        semester = qr_data.get("semester")

        df_students = pd.read_excel('students.xlsx')
        if str(roll) not in df_students['Roll No'].astype(str).values:
            return jsonify({"message": "\u274c Invalid student roll number"}), 400

        log_file = 'library_log.xlsx'
        if not os.path.exists(log_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Name', 'Roll No', 'Date', 'Class', 'Semester', 'Entry Time', 'Exit Time', 'Duration'])
            wb.save(log_file)

        wb = openpyxl.load_workbook(log_file)
        ws = wb.active

        today = datetime.now().date()
        now_time = datetime.now().strftime("%I:%M %p")

        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if str(row[1].value) == str(roll) and str(row[2].value) == str(today) and row[6].value is None:
                row[6].value = now_time
                entry_time = datetime.strptime(row[5].value, "%I:%M %p")
                exit_time = datetime.strptime(now_time, "%I:%M %p")
                minutes = (exit_time - entry_time).seconds // 60
                duration = f"{minutes // 60} hr {minutes % 60} min"
                row[7].value = duration
                message = f"\ud83d\udc4b Exit logged for {name} ({duration})"
                found = True
                break

        if not found:
            ws.append([name, roll, str(today), student_class, semester, now_time, None, None])
            message = f"\u2705 Entry logged for {name} at {now_time}"

        wb.save(log_file)
        return jsonify({"message": message})

    except Exception as e:
        print("Error:", e)
        return jsonify({"message": "\u274c Failed to process QR"}), 500

@app.route('/logs')
def show_logs():
    try:
        if not os.path.exists("library_log.xlsx"):
            return "<p>No records yet.</p>"

        df = pd.read_excel("library_log.xlsx")
        today = datetime.now().date()
        df_today = df[df['Date'] == str(today)]

        if df_today.empty:
            return "<p>No entries yet today.</p>"

        table_html = "<table border='1' cellpadding='5' style='margin:auto;border-collapse:collapse;'>"
        table_html += "<tr><th>Name</th><th>Roll</th><th>Entry</th><th>Exit</th><th>Duration</th></tr>"

        for _, row in df_today.iterrows():
            table_html += f"<tr><td>{row['Name']}</td><td>{row['Roll No']}</td><td>{row['Entry Time']}</td><td>{row['Exit Time'] or '-'}" \
                          f"</td><td>{row['Duration'] or '-'}</td></tr>"

        table_html += "</table>"
        return table_html

    except Exception as e:
        print("Error loading logs:", e)
        return "<p>Error loading log data</p>"

if __name__ == '__main__':
    app.run(debug=True)
